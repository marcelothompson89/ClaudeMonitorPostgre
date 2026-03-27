"""
Management command para importar eventos desde Template_Importacion_Eventos.xlsx
Formato esperado: Título* | Descripción | Fecha Inicio* | Fecha Fin | Tipo* | Ubicación
"""
import glob
import os
from datetime import date, datetime

from django.core.management.base import BaseCommand
from alertas.models import Evento


TIPOS_VALIDOS = {'evento', 'comite', 'politico', 'feriado', 'conferencia', 'otro'}


class Command(BaseCommand):
    help = 'Importa eventos desde archivo Excel con formato de template'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            type=str,
            help='Ruta al archivo Excel (por defecto busca Template_Importacion_Eventos.xlsx)',
        )
        parser.add_argument(
            '--limpiar',
            action='store_true',
            help='Eliminar todos los eventos existentes antes de importar',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR(
                'openpyxl no está instalado. Ejecute: pip install openpyxl'
            ))
            return

        # Buscar archivo Excel
        archivo = options.get('archivo')
        if not archivo:
            from django.conf import settings
            base_dir = settings.BASE_DIR
            # Primero buscar Template_Importacion_Eventos.xlsx
            candidato = os.path.join(str(base_dir), 'Template_Importacion_Eventos.xlsx')
            if os.path.exists(candidato):
                archivo = candidato
            else:
                # Fallback a CALENDARIO*.xlsx
                archivos = glob.glob(os.path.join(str(base_dir), 'CALENDARIO*.xlsx'))
                if not archivos:
                    self.stderr.write(self.style.ERROR(
                        'No se encontró Template_Importacion_Eventos.xlsx ni CALENDARIO*.xlsx'
                    ))
                    return
                archivo = archivos[0]

        nombre = os.path.basename(archivo)
        self.stdout.write(f'Leyendo archivo: {nombre}')

        wb = openpyxl.load_workbook(archivo)
        ws = wb[wb.sheetnames[0]]

        if options.get('limpiar'):
            count = Evento.objects.all().delete()[0]
            self.stdout.write(self.style.WARNING(f'Se eliminaron {count} eventos existentes.'))

        eventos_creados = 0
        eventos_existentes = 0
        errores = 0

        for row in range(2, ws.max_row + 1):
            titulo = ws.cell(row=row, column=1).value
            descripcion = ws.cell(row=row, column=2).value or ''
            fecha_inicio_raw = ws.cell(row=row, column=3).value
            fecha_fin_raw = ws.cell(row=row, column=4).value
            tipo_raw = ws.cell(row=row, column=5).value
            ubicacion = ws.cell(row=row, column=6).value or ''

            # Saltar filas vacías o separadores
            if not titulo or not fecha_inicio_raw:
                continue
            titulo = str(titulo).strip()
            if titulo.startswith('---') or titulo.lower().startswith('completar'):
                continue

            # Parsear fecha inicio
            fecha_inicio = self._parse_fecha(fecha_inicio_raw)
            if not fecha_inicio:
                self.stderr.write(self.style.WARNING(
                    f'Fila {row}: No se pudo parsear fecha_inicio "{fecha_inicio_raw}" - saltando'
                ))
                errores += 1
                continue

            # Parsear fecha fin (opcional)
            fecha_fin = self._parse_fecha(fecha_fin_raw) if fecha_fin_raw else None

            # Validar tipo
            tipo = str(tipo_raw).strip().lower() if tipo_raw else 'evento'
            # Mapear variantes comunes
            tipo = tipo.replace('é', 'e').replace('í', 'i')
            if tipo == 'comité':
                tipo = 'comite'
            elif tipo == 'político':
                tipo = 'politico'
            if tipo not in TIPOS_VALIDOS:
                self.stderr.write(self.style.WARNING(
                    f'Fila {row}: Tipo "{tipo_raw}" no válido, usando "evento"'
                ))
                tipo = 'evento'

            obj, created = Evento.objects.get_or_create(
                titulo=titulo,
                fecha_inicio=fecha_inicio,
                defaults={
                    'descripcion': str(descripcion).strip(),
                    'fecha_fin': fecha_fin,
                    'tipo': tipo,
                    'ubicacion': str(ubicacion).strip(),
                }
            )
            if created:
                eventos_creados += 1
                self.stdout.write(f'  + {titulo} ({fecha_inicio})')
            else:
                eventos_existentes += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nImportación completada: {eventos_creados} eventos creados, '
            f'{eventos_existentes} ya existían, {errores} errores.'
        ))

    def _parse_fecha(self, valor):
        """Convierte un valor de celda Excel a date."""
        if valor is None:
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        # Intentar parsear string con formatos comunes
        valor_str = str(valor).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(valor_str, fmt).date()
            except ValueError:
                continue
        return None
