"""
Script para crear el template de Excel para importación de eventos
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

def crear_template_eventos():
    """Crea un archivo Excel con el template para importar eventos"""

    wb = Workbook()

    # ===== HOJA 1: TEMPLATE DE DATOS =====
    ws_template = wb.active
    ws_template.title = "Template Eventos"

    # Encabezados
    headers = ["Título*", "Descripción", "Fecha Inicio*", "Fecha Fin", "Tipo*", "Ubicación"]
    ws_template.append(headers)

    # Formatear encabezados
    header_fill = PatternFill(start_color="002D62", end_color="002D62", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for col_num, header in enumerate(headers, 1):
        cell = ws_template.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fila de ejemplo
    ejemplo = [
        "ILAR Board of Directors",
        "Reunión trimestral del Board\nPARTICIPA: Juan, Melissa\nTemas: Planificación anual",
        date(2026, 6, 15),
        date(2026, 6, 16),
        "comite",
        "Virtual"
    ]
    ws_template.append(ejemplo)

    # Formatear fila de ejemplo con fondo gris claro
    ejemplo_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    for col_num in range(1, 7):
        cell = ws_template.cell(row=2, column=col_num)
        cell.fill = ejemplo_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Agregar 2 filas más de ejemplo adicionales
    ejemplos_adicionales = [
        [
            "Elecciones Presidenciales Colombia",
            "Primera vuelta elecciones presidenciales",
            date(2026, 5, 30),
            date(2026, 5, 31),
            "politico",
            "Colombia"
        ],
        [
            "FIFARMA Evento Anual",
            "Evento anual FIFARMA\nPARTICIPA: Juan\nTemas: Candidatos presidenciales Brasil",
            date(2026, 5, 5),
            date(2026, 5, 6),
            "conferencia",
            "Brasilia, Brasil"
        ]
    ]

    for ejemplo_row in ejemplos_adicionales:
        ws_template.append(ejemplo_row)
        for col_num in range(1, 7):
            cell = ws_template.cell(row=ws_template.max_row, column=col_num)
            cell.fill = ejemplo_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Agregar separador y espacio para datos del usuario
    ws_template.append([])  # Fila vacía
    ws_template.append(["--- COMPLETAR A PARTIR DE AQUÍ ---", "", "", "", "", ""])
    separador_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    separador_font = Font(bold=True, size=11)
    for col_num in range(1, 7):
        cell = ws_template.cell(row=ws_template.max_row, column=col_num)
        cell.fill = separador_fill
        cell.font = separador_font
        cell.alignment = Alignment(horizontal="center")

    # Agregar 20 filas vacías para que el usuario complete
    for _ in range(20):
        ws_template.append(["", "", "", "", "", ""])

    # Validación de datos para la columna "Tipo" (columna E)
    tipos_validos = '"evento,comite,politico,feriado,conferencia,otro"'
    dv = DataValidation(type="list", formula1=tipos_validos, allow_blank=False)
    dv.error = 'Por favor seleccione un tipo válido'
    dv.errorTitle = 'Entrada inválida'
    dv.prompt = 'Seleccione: evento, comite, politico, feriado, conferencia u otro'
    dv.promptTitle = 'Tipo de Evento'

    # Aplicar validación desde la fila 6 (después del separador) hasta la 100
    ws_template.add_data_validation(dv)
    dv.add(f"E6:E100")

    # Ajustar anchos de columna
    ws_template.column_dimensions['A'].width = 40  # Título
    ws_template.column_dimensions['B'].width = 50  # Descripción
    ws_template.column_dimensions['C'].width = 15  # Fecha Inicio
    ws_template.column_dimensions['D'].width = 15  # Fecha Fin
    ws_template.column_dimensions['E'].width = 15  # Tipo
    ws_template.column_dimensions['F'].width = 30  # Ubicación

    # Ajustar altura de filas de ejemplo
    ws_template.row_dimensions[1].height = 30
    ws_template.row_dimensions[2].height = 50
    ws_template.row_dimensions[3].height = 40
    ws_template.row_dimensions[4].height = 50

    # ===== HOJA 2: INSTRUCCIONES =====
    ws_instrucciones = wb.create_sheet("Instrucciones")

    # Título
    ws_instrucciones.append(["INSTRUCCIONES PARA COMPLETAR EL CALENDARIO DE EVENTOS"])
    ws_instrucciones.cell(row=1, column=1).font = Font(bold=True, size=16, color="002D62")
    ws_instrucciones.merge_cells('A1:D1')
    ws_instrucciones.append([])

    # Contenido de instrucciones
    instrucciones = [
        ["1. CAMPOS OBLIGATORIOS (marcados con *):"],
        ["   • Título*", "Nombre descriptivo del evento"],
        ["   • Fecha Inicio*", "Fecha en formato DD/MM/YYYY o formato fecha de Excel"],
        ["   • Tipo*", "Seleccionar de la lista desplegable (haga clic en la celda)"],
        [""],
        ["2. CAMPOS OPCIONALES:"],
        ["   • Descripción", "Agregar notas, participantes, temas a tratar (puede usar Alt+Enter para saltos de línea)"],
        ["   • Fecha Fin", "Solo si el evento dura más de un día"],
        ["   • Ubicación", "Ciudad, país o lugar del evento"],
        [""],
        ["3. TIPOS DE EVENTOS DISPONIBLES:"],
        ["   • evento", "Eventos generales de la organización"],
        ["   • comite", "Reuniones de comités (Regulatorio, Public Affairs, Suplementos, etc.)"],
        ["   • politico", "Eventos políticos (elecciones, cambios de gobierno, inauguraciones)"],
        ["   • feriado", "Días festivos, feriados nacionales, Semana Santa, Carnaval"],
        ["   • conferencia", "Congresos, seminarios, cumbres, foros, encuentros profesionales"],
        ["   • otro", "Cualquier evento que no encaje en las categorías anteriores"],
        [""],
        ["4. FORMATO DE FECHAS:"],
        ["   • Puede usar formato DD/MM/YYYY (ejemplo: 15/06/2026)"],
        ["   • O usar el selector de fechas de Excel (clic derecho > Formato de celdas > Fecha)"],
        ["   • Las fechas deben ser válidas y estar en el futuro o año actual"],
        [""],
        ["5. IMPORTANTE - EVITAR DUPLICADOS:"],
        ["   • La combinación de Título + Fecha Inicio debe ser única"],
        ["   • Si ya existe un evento con el mismo título y fecha, NO se creará duplicado"],
        ["   • Revise que no esté ingresando el mismo evento dos veces"],
        [""],
        ["6. RECOMENDACIONES:"],
        ["   • No usar caracteres especiales problemáticos en el Título: / \\ * ? < > |"],
        ["   • Para textos largos en Descripción, usar Alt+Enter para saltos de línea"],
        ["   • No dejar filas vacías entre eventos (saltear solo al finalizar)"],
        ["   • Revisar los ejemplos en la hoja 'Template Eventos' como guía"],
        [""],
        ["7. EJEMPLOS DE USO:"],
        [""],
        ["   Ejemplo 1 - Reunión de Comité:"],
        ["   Título:", "Comité Regulatorio"],
        ["   Descripción:", "Revisión de normativas Q1 2026\\nPARTICIPA: Juan, Melissa"],
        ["   Fecha Inicio:", "15/03/2026"],
        ["   Fecha Fin:", "(vacío si es de un solo día)"],
        ["   Tipo:", "comite"],
        ["   Ubicación:", "Virtual"],
        [""],
        ["   Ejemplo 2 - Conferencia Internacional:"],
        ["   Título:", "WHA 79.ª Asamblea Mundial de la Salud"],
        ["   Descripción:", "Asamblea anual OMS\\nPARTICIPA: Juan\\nInscripción antes del 1/5"],
        ["   Fecha Inicio:", "18/05/2026"],
        ["   Fecha Fin:", "23/05/2026"],
        ["   Tipo:", "conferencia"],
        ["   Ubicación:", "Ginebra, Suiza"],
        [""],
        ["   Ejemplo 3 - Evento Político:"],
        ["   Título:", "Elecciones Presidenciales Colombia (1ra vuelta)"],
        ["   Descripción:", "Primera vuelta elecciones presidenciales"],
        ["   Fecha Inicio:", "30/05/2026"],
        ["   Fecha Fin:", "30/05/2026"],
        ["   Tipo:", "politico"],
        ["   Ubicación:", "Colombia"],
        [""],
        ["8. DESPUÉS DE COMPLETAR:"],
        ["   • Guardar el archivo con un nombre descriptivo (ej: Eventos_ILAR_2026.xlsx)"],
        ["   • Enviar al administrador del sistema"],
        ["   • El administrador ejecutará: python manage.py import_eventos --archivo='ruta/archivo.xlsx'"],
        ["   • Opcionalmente con --limpiar para eliminar eventos existentes antes de importar"],
        [""],
        ["9. COLUMNAS DEL TEMPLATE:"],
        ["   Columna A - Título*", "(Obligatorio) - Máximo 500 caracteres"],
        ["   Columna B - Descripción", "(Opcional) - Texto libre, sin límite"],
        ["   Columna C - Fecha Inicio*", "(Obligatorio) - Formato fecha"],
        ["   Columna D - Fecha Fin", "(Opcional) - Formato fecha"],
        ["   Columna E - Tipo*", "(Obligatorio) - Lista desplegable con 6 opciones"],
        ["   Columna F - Ubicación", "(Opcional) - Máximo 300 caracteres"],
        [""],
        ["¿NECESITA AYUDA?"],
        ["Contacte al administrador del sistema para soporte técnico."],
    ]

    for row_data in instrucciones:
        ws_instrucciones.append(row_data)
        # Formatear según tipo de línea
        current_row = ws_instrucciones.max_row
        cell_a = ws_instrucciones.cell(row=current_row, column=1)

        if row_data[0].startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.")):
            cell_a.font = Font(bold=True, size=12, color="002D62")
        elif "Ejemplo" in row_data[0]:
            cell_a.font = Font(bold=True, size=11, color="0066CC")
        elif row_data[0].startswith("   •"):
            cell_a.font = Font(size=10)
        elif "NECESITA AYUDA" in row_data[0]:
            cell_a.font = Font(bold=True, size=13, color="DC3545")

        # Wrap text para todas las celdas
        for col_num in range(1, 5):
            ws_instrucciones.cell(row=current_row, column=col_num).alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    # Ajustar anchos de columna en instrucciones
    ws_instrucciones.column_dimensions['A'].width = 35
    ws_instrucciones.column_dimensions['B'].width = 60
    ws_instrucciones.column_dimensions['C'].width = 20
    ws_instrucciones.column_dimensions['D'].width = 20

    # Guardar archivo
    filename = "Template_Importacion_Eventos.xlsx"
    wb.save(filename)
    print(f"Template creado exitosamente: {filename}")
    print(f"Hoja 1: Template Eventos (con ejemplos y validaciones)")
    print(f"Hoja 2: Instrucciones detalladas")
    return filename

if __name__ == "__main__":
    crear_template_eventos()
